# coding: latin1
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from datetime import datetime
from inhibit_lib.mylib import *

# ex.: 'C:/Users/as0x/Desktop/Inhibit.csv'
file_csv = 'C:/Users/alanr/Desktop/Inhibit.csv'
# verifica se existe o arquivo
is_file(file_csv)

wb = Workbook()
ws = wb.active

ws.append(['DATA', 'HORA', 'AREA', 'TAG', 'BYPASS'])        # Nome das colunas
for row in csv.reader(open(file_csv)):
    # limpa o TAG para ficar mais apresentável
    ws.append([cell.replace('/HMI', '') for cell in row])


ws.delete_rows(2)  # Deleta índice de coluna original do CSV
ws.insert_rows(1)  # Insere linha do título

#clr_dict = load_color("color.json")
color = Color("color.json")

"""         FORMATAÇÃO DA  ÁREA  DOS  DADOS         """

area = "A3:" + ws.dimensions.split(':')[1]
# Define formato da fonte
font = Font(name='Times New Roman', sz=13)
# recuo à esquerda
alignm = Alignment(indent=1)
# expessura e cor das bordas
side = Side('thin', color.clr('steelblue'))
thin_border = Border(left=side, right=side, top=side, bottom=side)
range_format(ws, area, font, alignm, thin_border)
# Largura das colunas (fator de acordo com o tamanho da fonte)
col_width(ws, factor=1.35)

"""         FORMATAÇÃO DO  ÍNDICE  DAS  COLUNAS         """

# Formata fonte, tamanho, cor e negrito
font = Font('Arial', 12, True, color=color.clr('white'))
# Preenchimento, tipo e cor
fill = PatternFill('solid', color.clr('steelblue'))
range_format(ws, 'A2:E2', font, alignm, thin_border, fill)

"""         FORMATAÇÃO DA  ÁREA  DO  TÍTULO           """

ws.merge_cells('A1:E1')
ws['A1'].value = ' LISTA  DE  INIBIÇÕES  DO  ICSS  -  P55'
# Formata fonte, tamanho, cor e negrito
font = Font('Arial', 16, True, color=color.clr('navy'))
alignm = Alignment('center', 'center')
fill = PatternFill('solid', color.clr('whitesmoke'))
side = Side('double', color.clr('navy'))
_border = Border(top=side)
range_format(ws, 'A1:E1', font, alignm, _border, fill=fill)

# Altura da linha do título
ws.row_dimensions[1].height = 30

img = Image('logo.png')                                 # Gerar imagem logo BR
ws.add_image(img, 'A1')

"""         FINALIZAÇÃO           """

# Levanta as infomrmações de data e hora
now = datetime.now()
print_settings(ws)                                      # Ajuste da impressão
# Ajuste da visualização
view_settings(ws, now)

# ex.: 'C:/Users/as0x/Desktop/Inhibit_'
outfile = "Inhibit_"
# Arquivo com data e hora no nome
outfile += now.strftime("%Y%m%d%H%M") + ".xlsx"

# Grava o arquivo excel xlsx
write_file(wb, outfile, execute=False)
